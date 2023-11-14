from django.contrib.auth.decorators import login_required
from django.db.models import ExpressionWrapper, F, Max, Sum, fields
from django.http import Http404, HttpResponse
from django.shortcuts import redirect, render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side, fills

from laboratorio.models import Consumable, History, Permanent
from principal.decorators import allowed_users

material_roles = ["suporte", "material"]


@login_required(login_url="/dashboard/login")
@allowed_users(allowed_roles=material_roles)
def home_view(request):
    return redirect("consumable_dashboard")


@login_required(login_url="/dashboard/login")
@allowed_users(allowed_roles=material_roles)
def consumable_materials_view(request):
    return render(
        request,
        "laboratorio/dashboard/consumivel/dashboard.html",
    )


@login_required(login_url="/dashboard/login")
@allowed_users(allowed_roles=material_roles)
def update_consumable_view(request, pk):
    try:
        item = Consumable.objects.get(id=pk)

    except Consumable.DoesNotExist as err:
        raise Http404(err) from err

    context = {"item": item}
    return render(
        request,
        "laboratorio/dashboard/consumivel/editar.html",
        context,
    )


@login_required(login_url="/dashboard/login")
@allowed_users(allowed_roles=material_roles)
def materials_consumable_list_view(request):
    return render(
        request,
        "laboratorio/dashboard/consumivel/relacao.html",
    )


@login_required(login_url="/dashboard/login")
@allowed_users(allowed_roles=material_roles)
def permanent_materials_view(request):
    return render(
        request,
        "laboratorio/dashboard/permanente/dashboard.html",
    )


@login_required(login_url="/dashboard/login")
@allowed_users(allowed_roles=material_roles)
def update_permanent_view(request, pk):
    try:
        item = Permanent.objects.get(id=pk)

    except Permanent.DoesNotExist as err:
        raise Http404(err) from err

    context = {"item": item}
    return render(
        request,
        "laboratorio/dashboard/permanente/editar.html",
        context,
    )


@login_required(login_url="/dashboard/login")
@allowed_users(allowed_roles=material_roles)
def materials_permanent_list_view(request):
    return render(
        request,
        "laboratorio/dashboard/permanente/relacao.html",
    )


def highs_and_lows(queryset, measure_unit: str):
    queryset = queryset.annotate(
        diff=ExpressionWrapper(
            F("quantity") - F("prev_quantity"),
            output_field=fields.IntegerField(),
        ),
    )
    highs = (
        queryset.filter(diff__gt=0).aggregate(
            total_highs=Sum("diff"),
        )["total_highs"]
        or 0
    )
    lows = abs(
        queryset.filter(diff__lt=0).aggregate(
            total_lows=Sum("diff"),
        )["total_lows"]
        or 0,
    )

    if measure_unit:
        highs = f"{highs} {measure_unit.lower()}"
        lows = f"{lows} {measure_unit.lower()}"

    return [highs, lows]


@login_required(login_url="/dashboard/login")
@allowed_users(allowed_roles=material_roles)
def report_view(request, year):
    workbook = Workbook()

    ws = workbook.active
    ws.freeze_panes = ws["B4"]

    title = f"Relatório de entradas e saídas {year}"
    filename = f"relatorio_de_entradas_e_saídas_{year}.xlsx"
    operations = ["Entrada", "Saída"]
    columns = [
        "Janeiro",
        "Fevereiro",
        "Março",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
        f"{year}/Total",
    ]

    data = []
    materials = Consumable.objects.all()
    for material in materials:
        history = material.history.filter(created__year=year)
        if not history:
            continue

        material_report = [material.name]

        material_report.extend(
            value
            for month in range(1, 13)
            for value in highs_and_lows(
                history.filter(created__month=month),
                material.measure_unit,
            )
        )

        material_report += highs_and_lows(history, material.measure_unit)

        data.append(material_report)

    # TITLE
    cell = ws.cell(row=1, column=1, value=title.upper())
    cell.font = cell.font.copy(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28.35
    ws.merge_cells(
        start_row=1,
        end_row=1,
        start_column=1,
        end_column=(len(columns) * len(operations)) + 1,
    )

    # PRODUCT CELL
    cell = ws.cell(row=2, column=1, value="Produto".upper())
    cell.font = cell.font.copy(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[cell.column_letter].auto_size = True
    ws.merge_cells(
        start_row=2,
        end_row=3,
        start_column=1,
        end_column=1,
    )

    # MONTHS AND TOTAL CELLS
    for index, value in enumerate(columns, start=2):
        column = (index - 1) * len(operations) - (len(operations) - 2)

        cell = ws.cell(row=2, column=column, value=value.upper())
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = cell.font.copy(bold=True)
        ws.merge_cells(
            start_row=2,
            end_row=2,
            start_column=column,
            end_column=column + len(operations) - 1,
        )

    # 2ND LINE TABLE HEAD
    head = [op for _ in columns for op in operations]
    for column, value in enumerate(head, start=2):
        cell = ws.cell(row=3, column=column, value=value.upper())
        cell.font = cell.font.copy(bold=True)
        ws.column_dimensions[cell.column_letter].width = 12

    # PRODUCT DATA
    fill = PatternFill(
        start_color="FFFF00",
        end_color="FFFF00",
        fill_type=fills.FILL_SOLID,
    )

    for row, row_data in enumerate(data, start=4):
        for column, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=column, value=value)
            cell.fill = fill

    # BORDERS
    default_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = default_border

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    workbook.save(response)

    return response


@login_required(login_url="/dashboard/login")
@allowed_users(allowed_roles=material_roles)
def reports_view(request):
    context = {
        "reports": (
            History.objects.values("created__year")
            .annotate(last_modified=Max("modified"))
            .order_by("-created__year")
        ),
    }

    return render(
        request,
        "laboratorio/dashboard/consumivel/relatorios.html",
        context,
    )
